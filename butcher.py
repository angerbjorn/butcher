#!/usr/bin/python3
# -*- coding: utf-8 -*-
import sys
assert sys.version_info >= (3,0)

import time
import re
import os.path
import optparse
import xml.etree.ElementTree
import operator
import ipaddress
from xmljson import badgerfish as bf
import pymustache
import json
from xml.dom.minidom import parseString
import asn1crypto.x509
import ssl
import socket
socket.setdefaulttimeout(1)

mustache_template = """
<!DOCTYPE html>
<html lang="en">
    <head>
        <meta charset="utf-8">
        <title>Nessus report</title>
		<style>
@media screen {
			.sensitivity { float: left; background-color: #CC4E21; color: #ffffff; font-size: 16px; font-family:Tahoma; letter-spacing: 0px; border: 2px solid rgba(255, 255, 255, 0.5); border-radius: 10px; -moz-border-radius: 10px; overflow: hidden; padding: 2px 20px; text-align: center; text-shadow: 0px 0px 4px #737373; width: 60px; display: block; box-shadow:  0px 0px 4px 0px #A3A3A3; }
			.buttonCritical { background-color: FireBrick; }
			.buttonHigh { background-color: red; }
			.buttonMedium { background-color: DarkOrange; }
			.buttonLow { background-color: orange; }
			.buttonNone { background-color: gray; }
			.Critical { color: FireBrick; }
			.High { color: red; }
			.Medium { color: DarkOrange; }
			.Low { color: orange; }
			.None { color: gray; }
}
@media print {
			.sensitivity { float: left; color: #000000; font-size: 16px; font-family:Tahoma; overflow: hidden; text-align: center;  width: 30px; display: block; }
}
			.ReportItem {  padding-top 50px;  }
			.IP {padding-bottom: 0px !important; }
			.ReportItem div {white-space: pre-wrap; padding-bottom: 15px; }
			.desc::first-line {text-decoration: underline;}
			.pluginName {font-size: 120%;}
			.listItem {display:block; min-height:34px}
			.hosts {padding-left: 20px; padding-right: 20px; }
			hr {page-break-after: always;}
		</style>
	</head>
    <body>
        <h2>Nessus Report</h2>
        <div>Generated with the Butcher: https://github.com/angerbjorn/butcher</div>
		<h3>Table of findings:</h3>
		<p class="list">
			{{#severityList}}
			<div class="listItem" id="menu{{pluginID}}">
				<div class="sensitivity button{{risk_factor}}">{{risk_factor}}</div>
				<div class="pluginName"><span class="hosts">#{{count}}</span><a href="#item{{pluginID}}">{{pluginName}}</a></div>
			</div>
			{{/severityList}}
		</p>
		<hr>
		<h3>Details section:</h3>
		<p class="reports">
		
			<!--# ['@port', '@svc_name', '@protocol', '@severity', '@pluginID', '@pluginName', '@pluginFamily', 'agent', 'description', 'fname', 'plugin_modification_date', 'plugin_name', 'plugin_publication_date', 'plugin_type', 'risk_factor', 'script_version', 'solution', 'synopsis', 'plugin_output', 'IP', 'see_also', 'bid', 'cve', 'cvss3_base_score', 'cvss3_temporal_score', 'cvss3_temporal_vector', 'cvss3_vector', 'cvss_base_score', 'cvss_temporal_score', 'cvss_temporal_vector', 'cvss_vector', 'exploit_available', 'exploitability_ease', 'in_the_news', 'osvdb', 'vuln_publication_date', 'xref', 'cpe', 'patch_publication_date', 'cert', 'cwe', 'exploited_by_nessus', 'edb-id', 'icsa', 'cisco-bug-id', 'cisco-sa', 'iava', 'stig_severity', 'tra', 'zdi', 'canvas_package', 'exploit_framework_canvas', 'exploit_framework_core', 'attachment', 'exploit_framework_metasploit', 'metasploit_name', 'msft', 'unsupported_by_vendor', 'mskb', 'exploited_by_malware', 'mcafee-sb', 'cert-cc', 'iavb', 'vmsa', 'exploit_framework_exploithub', 'exploithub_sku', 'hp', 'rhsa', 'secunia', 'd2_elliot_name', 'exploit_framework_d2_elliot']-->
			{{#details}}
				{{#}}
				<div class="ReportItem" id="item{{@pluginID}}">
					<div class="pluginName" style=""><a href="#menu{{@pluginID}}"><span class="{{#risk_factor}}{{$}}{{/risk_factor}}">{{#risk_factor}}{{$}}{{/risk_factor}}</span> - {{@pluginName}}</a></div>

					{{#description}}<div class="description desc">Description:<br>{{$}}</div>{{/description}}
					{{#solution}}<div class="solution desc">Solution:<br>{{$}}</div>{{/solution}}
					{{#synopsis}}<div class="synopsis desc">Synopsis:<br>{{$}}</div>{{/synopsis}}
					<div class="plugin_output desc">Plugin output:<br>{{#plugin_output}}<br> &#11024; &#11024; &#11024; &#11024; &#11024; &#11024; Output for host {{IP}}{{#hostname}}/{{hostname}}{{/hostname}}&nbsp;:&nbsp;{{port}}&nbsp;({{svc_name}}) {{#location}}loc:&nbsp;{{location}}{{/location}} &#8628; &#8628; &#8628; &#8628; &#8628; &#8628; <br>{{output}}<br>{{/plugin_output}}</div>
					{{#see_also}}<div class="see_also desc">See also:<br>{{$}}</div>{{/see_also}}
					
					{{#exploit_available}}<div class="exploit_available">Exploit available: {{$}}</div>{{/exploit_available}}
					{{#exploit_framework_metasploit}}<div class="">exploit metasploit: {{$}}</div>{{/exploit_framework_metasploit}}
					{{#exploit_framework_core}}<div class="">exploit core: {{$}}</div>{{/exploit_framework_core}}
					{{#exploit_framework_exploithub}}<div class="">exploit exploithub: {{$}}</div>{{/exploit_framework_exploithub}}
					{{#exploit_framework_canvas}}<div class="">exploit canvas: {{$}}</div>{{/exploit_framework_canvas}}
					{{#exploit_framework_d2_elliot}}<div class="">exploit d2_elliot: {{$}}</div>{{/exploit_framework_d2_elliot}}
					{{#secunia}}<div class="">secunia:{{$}}</div>{{/secunia}}
					{{#attachment}}<div class="">attachment:{{$}}</div>{{/attachment}}
					{{#in_the_news}}<div class="">in_the_news:{{$}}</div>{{/in_the_news}}
					
					<div class="IPS">Affected hosts: {{#IP}}<div class="IP">{{IP}}{{#hostname}}/{{hostname}}{{/hostname}} : {{port}}&nbsp;({{svc_name}}) {{#location}} - Location: {{location}}{{/location}}  Hostname-probability:<i>{{hostNameProbability}}</i></div>{{/IP}}</div>
					<div>Plugin ID: {{@pluginID}}</div>
					<div class="bid ">bid: {{#bid}}{{$}} {{/bid}}</div>
					<div class="xref ">xref: {{#xref}}{{$}} {{/xref}}</div>
					<div class="cve ">cve: {{#cve}}{{$}} {{/cve}}</div>
					<div class="osvdb ">osvdb: {{#osvdb}}{{$}} {{/osvdb}}</div>

				</div>
				<hr>
				{{/}}
			{{/details}}
		</p>
    </body>
</html>
"""
searchDomain = False 
lastDomain = None 

def tag( rh, key ):
	# hostname is stored as <ReportHost><HostProperties><tag name="host-fqdn">v-something.somedomain.de</tag> not all hosts has a name...
	dn = rh.find(".HostProperties//*[@name='"+key+"']")
	if type(dn) == xml.etree.ElementTree.Element:
		return dn.text.lower()
	return ""

def baseName( fqdn):
	global searchDomain, lastDomain
	if type(fqdn) != str:
		return fqdn
	r =  fqdn.split('.')
	if len(r) > 1:
		if not searchDomain:
			lastDomain = '.'.join(r[1:])
		return r[0]
	return fqdn


windows_ten_build_ver = {
	10240: '1507',
	10586: '1511',
	14393: '1607',
	15063: '1703',
	16299: '1709',
	17134: '1803',
	17763: '1809',
	18362: '1903',
	18363: '1909',
	19041: '2004',
	19042: '20H2',
	19043: '21H1'
}

def osNameHarmanizer( osName ) -> str : # SyntaxError on this line means python2 is used instead of python3 ...
	# enterprise is a company, business,
	# entreprise archaic form of enterprise
	# fix windows !!!!
	name = osName.strip().lower().replace('entreprise','enterprise').replace('®','').replace('(r)','').replace('service pack', 'sp').replace('microsoft windows', 'windows').replace('cisco ios cisco ios software', 'cisco ios')
	return name

def getBestOS( ReportHost ) -> str :

	# if a WMI report of OS exists, this looks more accurate :)
	# <ReportItem port="445" svc_name="cifs" protocol="tcp" severity="0" pluginID="24269" pluginName="WMI Available" pluginFamily="Windows">
	# ...
	# <plugin_output>The remote host returned the following caption from Win32_OperatingSystem:
	#
	#         Microsoft Windows 10 Enterprise</plugin_output>
	os = ReportHost.find('.//*[@pluginID="24269"]/plugin_output')
	if type(os) == xml.etree.ElementTree.Element:
		win = os.text.lower().split('\n')[2].strip()

		# <ReportItem port="445" svc_name="cifs" protocol="tcp" severity="0" pluginID="48942" pluginName="Microsoft Windows SMB Registry : OS Version and Processor Architecture" pluginFamily="Windows">
		# ...
		# <plugin_output>Operating system version = 10.18363
		# Architecture = x64
		# Build lab extended = 18362.1.amd64fre.19h1_release.190318-1202
		# </plugin_output>

		os = ReportHost.find('.//*[@pluginID="48942"]/plugin_output')
		if type(os) == xml.etree.ElementTree.Element:
			bo = os.text.lower().split('\n')[0].strip()
			if bo.startswith('operating system version = 10.'):
				build = int(bo.split('= 10.')[1])
				if windows_ten_build_ver.get( build ):
					return "%s ver %s build %s" %(osNameHarmanizer(win), windows_ten_build_ver[ build ], build )
				return "%s build %s" %(osNameHarmanizer(win), build )
		return osNameHarmanizer(win)

	os = tag(ReportHost, 'operating-system')
	if os:
		if os.startswith('microsoft windows server 2003\nmicrosoft windows vista\n'):
			return osNameHarmanizer('microsoft windows something')
		return osNameHarmanizer(os.replace('\n', ' or '))

	os = tag(ReportHost, 'cpe-1')
	if os:
		return osNameHarmanizer("Maybe: "+os.replace('\n', ' or '))

	os = tag(ReportHost, 'cpe-0')
	if os:
		return osNameHarmanizer("Maybe: "+os.replace('\n', ' or '))

	os = tag(ReportHost, 'os')
	if os:
		return osNameHarmanizer(os.replace('\n', ' or '))

	return 'unknown-os'

def getBestHostname( IP=False, ReportHost=False, no_lookups=False, no_scanning=False, resolvHelper=False):
	# find a name that match the IP, preferably netbios-name
	global searchDomain, lastDomain
	
	# 'host-ip','netbios-name', 'host-fqdn','host-rdns', 'hostname.wmi-domain'
	if not IP:
		IP = tag( ReportHost, 'host-ip')

	if resolvHelper:
		if IP in resolvHelper:
			return( resolvHelper[IP])

	if ReportHost:
		if not no_lookups:
			# try Netbios from Nessus first
			name = tag( ReportHost, 'netbios-name')
			if name and (resolve( name ) == IP ):
				return ( name, 'Correct: nessus-netbios')

			# fqdn from nessus, first as base
			name = baseName( tag( ReportHost, 'host-fqdn'))
			if name and (resolve( name ) == IP ):
				if not searchDomain:
					# check if we can find a correct searchDomain using this hostname
					if resolve( name + '.' + lastDomain ) == IP:
						searchDomain = lastDomain
						print( 'DNS Suffix found as %s and %s both resolves to %s' %(name, name + '.' + lastDomain, IP), file=sys.stderr)
				return ( name, 'Correct: nessus-fqdn as hostname only')

			# fqdn from nessus as is
			name = tag( ReportHost, 'host-fqdn')
			if name and (resolve( name ) == IP ):
				return ( name, 'Correct: nessus-fqdn as is')

			# RDNS
			name = tag( ReportHost, 'host-rdns')
			if name and not ( isIP( name ) ) and (resolve( name ) == IP ):
				return ( name, 'Correct: nessus-rdns as is')

			# nessus hostname
			name = tag( ReportHost, 'hostname')
			if name and (resolve( name ) == IP ):
				return ( name, 'Correct: nessus-hostname as is')

			# nessus hostname and wmi-domain
			name = tag( ReportHost, 'hostname') + '.' + tag( ReportHost, 'wmi-domain')
			if resolve( name ) == IP :
				return ( name, 'Correct: nessus-hostname and wmi-domain')

			# try reverse DNS full
			name = baseName( reverse( IP ))
			if name and (resolve( name ) == IP ):
				return ( name, 'Correct: reverse DNS lookup as hostname only')

			# try reverse DNS full
			name = reverse( IP )
			if name and (resolve( name ) == IP ):
				return ( name, 'Correct: reverse DNS lookup as fqdn')

			# grab from RDP CN

			if not no_scanning and tag( ReportHost, 'operating-system').lower().find('windows') != -1 :
				name = getCN( IP )
				if name:
					if baseName(resolve( name )) == IP :
						return ( baseName(name.lower()), 'Correct: RDP CN hostname only')
					elif resolve( name ) == IP :
						return ( name.lower(), 'Correct: RDP CN FQDN')
					elif ops.verbose:
						print( 'RDP CN name mismatch for ip', IP, 'CN', name, file=sys.stderr)

		##### nothing resolves to the correct name .... 
		
		# try Netbios from Nessus first 
		name = tag( ReportHost, 'netbios-name')
		if name:
			return ( name, 'Guess: unresolved nessus-netbios') 

		# fqdn from nessus as is 
		name = tag( ReportHost, 'host-fqdn')
		if name: 
			if searchDomain and name.endswith(searchDomain):
				return ( baseName(name), 'Guess: unresolved nessus-fqdn as hostname only')
			else:
				return ( name, 'Guess: unresolved nessus-fqdn, suffix mismatch')
			
				
		# hostname 
		name = tag( ReportHost, 'hostname') + '.' + tag( ReportHost, 'wmi-domain')
		if name != '.':
			return ( name, 'Guess: unresolved nessus-hostname and wmi-domain')

		
		# nothing worked 
		return ( '', 'Failure')
			
		
def resolve( hostname ):
	try:
		return socket.gethostbyname(hostname)
	except (socket.error, UnicodeError):
		return False
		
def reverse( ip ):
	try:
		ghba =  socket.gethostbyaddr(ip)
		if type(ghba) == tuple:
			return ghba[0].lower() 
	except (socket.error, UnicodeError):
		return False

def isIP( hostname ):
	try: 
		ipaddress.ip_address( hostname )
		return True
	except:
		return False
		
def getCNOpenSSL( ip , port=3389):
	try:
		cert = ssl.get_server_certificate((ip, port ))
		x509 = OpenSSL.crypto.load_certificate(OpenSSL.crypto.FILETYPE_PEM, cert)
		t = x509.get_subject().get_components()
		# [(b'CN', b'RWWBCH02.something.local')]
		if type( t ) == list:
			return t[0][1].decode('utf-8')
		return False
	except (ConnectionRefusedError, TimeoutError, socket.timeout, ConnectionResetError):
		return False

def getCN( ip , port=3389):
	try:
		cert = ssl.get_server_certificate((ip, port))
		c = asn1crypto.x509.Certificate.load( ssl.PEM_cert_to_DER_cert(cert) )
		return(c.subject.human_friendly[len('Common Name: '):])
	except (ConnectionRefusedError, TimeoutError, socket.timeout, ConnectionResetError, socket.gaierror, OSError):
		return False

def addPort( hostName, port ):
	if hostName == '':
		return ''
	if port == '0':
		return hostName
	return hostName + ':' + port 	
	

def getValue( key, data ):
	if data.get(key):
		return data.get(key)
	elif data.find(key) != None :
		return data.find(key).text
	else:
		print('Fatal: Unknown key error: "%s"\n\nSome common keys are:\n%s\nThe full nessus_v2 file format is documented in the nessus_v2_file_format.pdf paper.\nAlso, --format xml or json can be helpful to understand the data structure and keys used.' %(key, ' '.join(['port', 'svc_name', 'protocol', 'severity', 'pluginID', 'pluginName', 'pluginFamily', 'agent', 'description', 'fname', 'plugin_modification_date', 'plugin_name', 'plugin_publication_date', 'plugin_type', 'risk_factor', 'script_version', 'solution', 'synopsis', 'plugin_output', 'IP', 'see_also', 'bid', 'cve', 'cvss3_base_score', 'cvss3_temporal_score', 'cvss3_temporal_vector', 'cvss3_vector', 'cvss_base_score', 'cvss_temporal_score', 'cvss_temporal_vector', 'cvss_vector', 'exploit_available', 'exploitability_ease', 'in_the_news', 'osvdb', 'vuln_publication_date', 'xref', 'cpe', 'patch_publication_date', 'cert', 'cwe', 'exploited_by_nessus', 'edb-id', 'icsa', 'cisco-bug-id', 'cisco-sa', 'iava', 'stig_severity', 'tra', 'zdi', 'canvas_package', 'exploit_framework_canvas', 'exploit_framework_core', 'attachment', 'exploit_framework_metasploit', 'metasploit_name', 'msft', 'unsupported_by_vendor', 'mskb', 'exploited_by_malware', 'mcafee-sb', 'cert-cc', 'iavb', 'vmsa', 'exploit_framework_exploithub', 'exploithub_sku', 'hp', 'rhsa', 'secunia', 'd2_elliot_name', 'exploit_framework_d2_elliot'])), file=sys.stderr)
		exit()

def ipOrName( host: str ) -> str:
	# host can be either CIDR/IP-range, IP-address, hostname or comment
	if host.startswith('#'):
		return 'comment'
	if host.find(' ') != -1:
		print("Typo in hostname? There is a space in '%s'" % host, file=sys.stderr)
	try:
		socket.inet_aton( host )
		return('IP')
	except(socket.error):
		try:
			ipaddress.ip_network( host )
			return 'range'
		except(ValueError):
			if re.search("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$", host ):
				print( "Typo in IP address? %s is classed as a hostname" %host , file=sys.stderr)
			if re.search("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}/\d{1,2}$", host ):
				print( "Typo in CIDR? %s is classed as a hostname" %host , file=sys.stderr)
			if not re.search("^[a-z0-9.-]+$", host):
				print("Typo in hostname? '%s' is classed as a hostname" % host, file=sys.stderr)
	return 'hostname'

def dictConcat( dictionary :dict , key:str, value:str ) -> None:
	if key in dictionary:
		dictionary[ key ] =  "%s / %s" %( dictionary[ key ], value )
	else:
		dictionary[ key ] =  value

def PrettyTableWrapper(vertical_char=' '):
	imported = True
	try:
		from prettytable import PrettyTable
		from prettytable import MSWORD_FRIENDLY
	except ModuleNotFoundError:
		imported = False

	if imported:
		pt = PrettyTable()
		pt.set_style(MSWORD_FRIENDLY)
		pt.vertical_char = vertical_char
		pt.left_padding_width = 0
		pt.right_padding_width = 0
		return pt
	else:
		# output without prettytable, using tabs instead
		print( "Please install the prettytable python package to get straight column output", file=sys.stderr)
		class PrettyTable:
			def __init__(self):
				self.table = []
				self.field_names = []

			def add_row(self, *columns):
				self.table.append(columns[0])

			def __str__(self):
				s = '%s\n' % '\t'.join(self.field_names)
				for r in self.table:
					for w in r:
						s += str(w) + '\t'
					s += '\n'
				return s
			def get_string(self):
				return self.__str__()
		return PrettyTable()

def cachePath( report_task_id ):
	# create a path to a cache file in current running butchers base directory.
	return os.path.join(os.path.dirname( os.path.realpath(__file__) ), ".namecache.%s.json" %report_task_id )

#def cacheAge( filePath , expire=60*60*24*7):
# REMOVED as an old cache is most likely more correct !
#	# check if exits, if file and last modified
#	if not os.path.isfile( filePath ):
#		return  "No previous cache exists"
#	if time.time() - os.path.getmtime( filePath ) > expire:
#		return "Expired"
#	return "Fresh"

def cacheLoad( filePath ):
	if not os.path.isfile( filePath ):
		return  False, "No previous cache exists"
	f = open( filePath, 'r' )
	try:
		return True, json.loads( f.read() )
	except:
		return False, "Failed to load json data."
	finally:
		f.close()

def getReportTaskId(root, nessus_file):
	# the namechache need the report_task_id, or will use the filename as idenfifier for the local name cache
	# ...
	# <ServerPreferences>
	# ...
	# <preference><name>report_task_id</name>
	# <value>60933064-7047-e6d8-148e-c3915cefe8db6e87a8e0d90d7af1</value>
	for sp in root.iter('ServerPreferences'):
		for p in sp.findall("preference/[name='report_task_id']"):
			return p.find('value').text
	return os.path.splitext(os.path.basename(nessus_file))[0]


if __name__ == "__main__":
	parser = optparse.OptionParser(usage="Usage: %prog [OPTION]... <NESSUS_FILE>...", description="Compiles a report from one or more .nessus v2 files. Output can be text, html or excel. Filters can be set to text matches, severity, hosts, IP-networks, or nessus-IDs", epilog='Open Source MIT License. Written by Christian Angerbjorn')
	parser.add_option("-v", "--verbose", action="store_true", help='verbose output')
	parser.add_option("", "--force-glob", action="store_true", help='Expand wildcards such as * or ? in nessus file names, this is automatic in Windows cmd shell, but not in unix shells since they do that already')
	parser.add_option("", "--no-glob", action="store_true", help='Do NOT expand wildcards')

	group = optparse.OptionGroup(parser, "Output options, a few format and styles exists")
	group.add_option("-f", "--format",  default='text', help='Optional output format, either of [text, html, excel, json, xml, grep] Defaults to text.')
	# rename to overview, details and findings
	group.add_option("-O", "--style", help="Text output can be either: one line per finding (compact), one line per IP per finding (long), or one line host (host). Defaults to host")
	group.add_option("-o", "--output-file", help="Optional output file to save result as. Mandatory for Excel output")
	group.add_option("-K", "--output-sheet", default='', help="Excel sheet name to use when saving data")
	group.add_option("-H", "--html-template", help="Optional Mustache HTML template to use. As a starting point, see the mustache_template=  in this source code")
	group.add_option("-l", "--long", action="store_true", help="Deprecated, replaced with --style long")
	group.add_option("", "--vertical-char", default=' ', help="Use this character as vertical separator in text output. Defaults to space")
	parser.add_option_group(group)

	group = optparse.OptionGroup(parser, "Grep format options, This is a unix 'grep' or Windows 'find' friendly format, each line starts with IP and hostname")
	# add OptionGroup for grep
	group.add_option("-Q", "--grep-delim", default='|', help="Use this delimiter before and after each line output")
	group.add_option("-V", "--grep-raw", action="store_true", help="Hide host:port data on each line, useful to for example sample sets")
	group.add_option("-J", "--grep-plugin", action="store_true", help="Show plugin_output data only")
	group.add_option("-w", "--grep-description", action="store_true", help="Show description data only")
	parser.add_option_group(group)

	group = optparse.OptionGroup(parser, "Lookups mean active identification of routable hostnames. The result is automatically cached.")
	parser.add_option("-a", "--lookups", action="store_true", help='Do active name lookups when identifying hostname. The result will be cached.')
	parser.add_option("", "--lookups-no-certs", default=False, action="store_true", help='Disable remote host cerfiticate name grabbing - used when all other means of obtaining a hostname has failed as part of lookups. Require --lookups ')
	parser.add_option("", "--no-cache", action="store_true", help='Do not read from the name cache')
	parser.add_option("", "--save-cache", action="store_true", help='Update the cache without using --lookups that will do this automatically.')
	parser.add_option_group(group)

	group = optparse.OptionGroup(parser, "Add a hostname mapper that is used first to find a hostname.")
	# rename or add to cache-something
	group.add_option("-P", "--hostname-excel", action="append", default=[], help="Read hostname,ip and probability data cached from this excel spreadsheet, for example a previous --style host report. Use multiple times as needed.")
	group.add_option("-p", "--hostname-sheet", default='1', help="Sheet-name or index number (starting with 1) with --hostname-excel  Defaults to 1")
	group.add_option("-u", "--hostname-ip-column", default='H', help="Column name for 'IP-address' data with --hostname-excel  Defaults to H")
	group.add_option("-U", "--hostname-name-column", default='G', help="Column name for 'hostname' data with --hostname-excel  Defaults to G")
	group.add_option("-Y", "--hostname-probability-column", default='K', help="Column name for 'hostname' data with --hostname-excel  Defaults to K")
	parser.add_option_group(group)

	group = optparse.OptionGroup(parser, "Add supporting \"Location\" data to each host matched with either CIRD/IP-range, IP-address or hostname. Each host can show multiple matches")
	# rename or add location-tag
	group.add_option("-L", "--location-excel", action="append", default=[], help="Read Location data from this excel spreadsheet, host in --subnet-column and  --location-column for any data to show as Location. Use multiple times as needed.")
	group.add_option("-Z", "--location-sheet", default='1', help="Sheet-name or index number (starting with 1) with --location-excel  Defaults to 1")
	group.add_option("-C", "--location-host-column", default='A', help="Column name for 'hosts' data with --location-excel  Defaults to A")
	group.add_option("-X", "--location-column", default='B', help="Column name for 'location' data with --location-excel  Defaults to B")
	parser.add_option_group(group)

	group = optparse.OptionGroup(parser, 'Filter options, only show data that match criteria, or remove data that does not match criteria. Hosts can be either of IP address (IPv4), hostname or CIDR/range such as 10.0.0.0/24')
	group.add_option("-s", "--min-severity", default='0', help="Either none, low, medium, high, critical or a number from 0-4, where 0=None, 1=Low, 2=Medium, 3=High, and 4=Critical.")
	group.add_option("-S", "--max-severity", default='4', help="Either none, low, medium, high, critical or a number from 0-4, where 0=None, 1=Low, 2=Medium, 3=High, and 4=Critical.")

	group.add_option("-t", "--hosts", action="append", default=[], help="Show only these hosts in the report.  Use multiple times as needed.")
	group.add_option("-T", "--no-hosts", action="append", default=[], help="Exclude host from the report. Use multiple times as needed.")
	group.add_option("-r", "--hosts-file", action="append", default=[], help="Show only hosts from file, one per line. Use multiple times as needed.")
	group.add_option("-R", "--no-hosts-file", action="append", default=[], help="Exclude hosts from file, one per line. Use multiple times as needed.")
	group.add_option("-e", "--hosts-excel", action="append", default=[], help="Show only hosts from an excel spreadsheet. (Location data is also added similar to --location-excel) Use multiple times as needed.")
	group.add_option("-E", "--no-hosts-excel", action="append", default=[], help="Exclude hosts from an excel spreadsheet. Use multiple times as needed.")
	group.add_option("-z", "--excel-sheet", default='1', help="Sheet-name or index number (starting with 1) with --hosts-excel.  Defaults to 1")
	group.add_option("-c", "--excel-name-column", default='A', help="Column name for 'hosts' data with --hosts-excel  Defaults to A")
	group.add_option("-x", "--excel-location-column", default='B', help="Column name for 'location' data with --hosts-excel Defaults to B")

	group.add_option("-m", "--match", help="Only show results with a pluginName matching this regex search term, for example 'ms17-010'")
	group.add_option("-M", "--no-match", help="Only show results with a pluginName NOT matching this regex search term.")
	group.add_option("-k", "--match-key", default='pluginName', help="Use this key for --match, defaults to 'pluginName'")

	group.add_option("-i", "--id", action="append", default=[], help="Include only finding with this nessus ID the report. Use multiple times as needed.")
	group.add_option("-I", "--no-id", action="append", default=[], help="Exclude findings with this nessus ID the report. Use multiple times as needed.")
	group.add_option("-b", "--id-file", action="append", default=[], help="Include only finding, one per line, with this nessus ID the report. Use multiple times as needed.")
	group.add_option("-B", "--no-id-file", action="append", default=[], help="Exclude findings, one per line, with this nessus ID the report. Use multiple times as needed.")
	parser.add_option_group(group)

	group = optparse.OptionGroup(parser, "Singe purpose and special modes, does what they says and exits")
	group.add_option("-D", "--dump-xml-key", help='Json dump values corresponding to supplied key, then exit. Example key: preference')
	group.add_option("-d", "--dump-targets",  action="store_true", help='Compare targets used in the scans with a list from STDIN and exit')
	group.add_option("-W", "--targets-excel", help='Compare targets/IP ranges used in the scans with a list from this excel document and exit')
	parser.add_option_group(group)

	(ops, args) = parser.parse_args()
	if len(args) == 0:
		parser.error("At least one .nessus is required!")
		
	if (ops.force_glob or sys.platform.startswith('win')) and not ops.no_glob:
		import glob
		newArgs = []
		for gf in args:
			for f in glob.glob(gf):
				newArgs.append(f)
		args = newArgs

	severity = {'none': 0, 'low': 1, 'medium': 2, 'high': 3, 'critical': 4}
	# min
	if ops.min_severity:
		if ops.min_severity.isdigit():
			ops.min_severity = int(ops.min_severity)
		else:
			try:
				ops.min_severity = severity[ops.min_severity.lower()]
			except: 
				parser.error('--min-severity has to be either a number from 0-4 or none, low, medium, high, critical')
	# max
	if ops.max_severity:
		if ops.max_severity.isdigit():
			ops.max_severity = int(ops.max_severity)
		else:
			try:
				ops.max_severity = severity[ops.max_severity.lower()]
			except:
				parser.error('--max-severity has to be either a number from 0-4 or none, low, medium, high, critical')

	if ops.format and ops.format not in ['text', 'html', 'excel', 'json', 'xml', 'grep']:
		parser.error("Format can only be one of [text, html, excel, json, xml grep]")
		
	if ops.long and ops.style :
		parser.error('--long is deprecated and replaced with --style long')
	if ops.long and not ops.style:
		ops.style = 'long'
		
	if ops.style and ops.style not in ['compact', 'long', 'host']:
		parser.error("--style can only be one of [compact, long, host]")
	
	# set the default value after --long check 
	if not ops.style:
		ops.style = 'host'
	
	if ops.format == 'excel' and not ops.output_file:
		parser.error('Excel output require an --output-file')
		
	# read in/exclude nessus IDs from file 
	for nf in ops.id_file:
		with open( nf, 'r' , encoding="utf8") as f: 
			for n in f.read().splitlines():
				ops.id.append( n )
	for nf in ops.no_id_file:
		with open( nf, 'r' , encoding="utf8") as f: 
			for n in f.read().splitlines():
				ops.no_id.append( n )

	# read in/exclude hosts from file
	for nf in ops.hosts_file:
		with open( nf, 'r' , encoding="utf8") as f: 
			for n in f.read().splitlines():
				ops.hosts.append( n )
	for nf in ops.no_hosts_file:
		with open( nf, 'r' , encoding="utf8") as f: 
			for n in f.read().splitlines():
				ops.no_hosts.append( n )

	# read include hosts from excel
	if ops.hosts_excel or ops.no_hosts_excel or ops.format == 'excel' or ops.location_excel or ops.hostname_excel:
		from openpyxl import load_workbook, Workbook

	# hostname resolv helper
	resolvHelper = {'127.0.0.1':('localhost', 'Correct')}
	
	for nessus_file in args:
		report_task_id = False
		with open(nessus_file, 'r', encoding="utf8") as f:
			root = xml.etree.ElementTree.parse(f).getroot()
			report_task_id = getReportTaskId(root, nessus_file)
			
			cFile = cachePath(report_task_id)
			if not ops.lookups and not ops.no_cache:
				cJson = cacheLoad(cFile)
				if not cJson[0]:
					print("Warning: No name lookup cache for %s, consider active name lookups using --lookups" % os.path.basename(nessus_file), file=sys.stderr)
				if cJson[0]:
					if ops.verbose:
						print("Using name lookup cache for %s report_task_id %s" % (os.path.basename(nessus_file), report_task_id), file=sys.stderr)
					resolvHelper.update(cJson[1])
		# namecache done.
	
	for ef in ops.hostname_excel:
		if not os.path.isfile( ef ):
			print("the --hostname-excel %s file is missing! (but will continue anyway)" %ef , file=sys.stderr)
			break
		originalSheet = ops.hostname_sheet
		wb = load_workbook( filename=ef )
		if ops.hostname_sheet.isdigit():
			ops.hostname_sheet = wb.sheetnames[int(ops.hostname_sheet)-1]
		for i, row in enumerate(wb[ ops.hostname_sheet ].iter_rows()):
			IP  =  row[ ord(ops.hostname_ip_column.upper())-65 ].value
			if IP and not IP.startswith('#'):
				hname = row[ ord(ops.hostname_name_column.upper())-65 ].value
				if hname:
					hname = hname.lower()
					if hname.startswith('≈'):
						hname = hname[1:]
				else:
					hname = ""
				proba = row[ ord(ops.hostname_probability_column.upper())-65 ].value
				resolvHelper[ IP ] = (hname, proba)
		ops.hostname_sheet = originalSheet

	includeHosts = set() # list of IP-addresses or hostnames. (hence not ipaddress.overlap used...)
	excludeHosts = set()
	netLookup = {} # subnet: location-data
	hostLookup = {} # hostname or IP : location-data
	for ef in ops.hosts_excel:
		originalSheet = ops.excel_sheet
		wb = load_workbook( filename=ef )
		if ops.excel_sheet.isdigit():
			ops.excel_sheet = wb.sheetnames[int(ops.excel_sheet)-1]
		for i, row in enumerate(wb[ ops.excel_sheet ].iter_rows()):
			host  =  row[ ord(ops.excel_name_column.upper())-65 ].value
			if not host:
				continue
			host = host.lower()
			location = row[ ord(ops.excel_location_column.upper())-65 ].value
			hostType = ipOrName( host )

			if hostType == 'range':
				if host not in netLookup:
					ops.hosts.append( host )
				dictConcat( netLookup, host, location )

			if hostType == 'IP' or hostType == 'hostname':
				includeHosts.add( host )
				dictConcat( hostLookup, host, location)
		ops.excel_sheet = originalSheet

	# location data only
	for ef in ops.location_excel:
		originalSheet = ops.location_sheet
		wb = load_workbook( filename=ef )
		if ops.location_sheet.isdigit():
			ops.location_sheet = wb.sheetnames[int(ops.location_sheet)-1]
		for i, row in enumerate(wb[ ops.location_sheet ].iter_rows()):
			host  =  row[ ord(ops.location_host_column.upper())-65 ].value.lower()
			location = row[ ord(ops.location_column.upper())-65 ].value
			hostType = ipOrName( host )
			if hostType == 'range':
				for addr in ipaddress.ip_network( host ):
					dictConcat( hostLookup, str(addr), location)
			if hostType == 'IP' or hostType == 'hostname':
				dictConcat(hostLookup, host, location)
		ops.location_sheet = originalSheet

	for ef in ops.no_hosts_excel:
		originalSheet = ops.excel_sheet
		wb = load_workbook( filename=ef )
		if ops.excel_sheet.isdigit():
			ops.excel_sheet = wb.sheetnames[int(ops.excel_sheet)-1]
		for i, row in enumerate(wb[ ops.excel_sheet ].iter_rows()):
			ops.no_hosts.append( row[ ord(ops.excel_name_column.upper())-65 ].value )
		ops.excel_sheet = originalSheet

	if ops.verbose:
		for k in ops.id:
			print("Include nessus ID: %s" %k, file=sys.stderr)
		for k in ops.no_id:
			print("Exclude nessus ID: %s" %k, file=sys.stderr)

	for n in ops.hosts:
		n = n.lower()
		if ops.verbose:
			print("Including hosts %s" %n, file=sys.stderr)
		dataType = ipOrName( n )
		if dataType == 'range':
			for addr in ipaddress.ip_network(n):
				includeHosts.add( str(addr) )
				# add host lookup data to each host
				if n in netLookup:
					dictConcat( hostLookup, str(addr), netLookup[ n ])
		if dataType == 'hostname' or dataType == 'IP':
			includeHosts.add( n )

	for n in ops.no_hosts:
		n = n.lower()
		if ops.verbose:
			print("Excluding hosts %s" %n, file=sys.stderr)
		hostType = ipOrName( n )
		if hostType == 'range':
			for addr in ipaddress.ip_network(n):
				excludeHosts.add( str(addr) )
		if hostType == 'hostname' or hostType == 'IP':
			excludeHosts.add( n )

	# dict of pluginId and finding data 
	compact = {} # {"pluginID, meaning 12345":{findings-data}}
	# list of some findidng data, so that we can sort in order or severity 
	findings = [] # [{findings-data}, {findings-data}, {findings-data}]
	long_findings = [] # [{findings-data}, {findings-data}, {findings-data}]
	host_stats = { "location":"", "displayAddr":"", "displayOs":"",'none': 0, 'low': 0, 'medium': 0, "high":0, "critical":0 }
	host_findings = {} # {IP:host_stats}
	sum_findings = host_stats.copy()
	locationMatch = False
	
	if ops.dump_targets or ops.targets_excel:
		nessusNetworks = set()
		inNetworks = set()
		
		if ops.targets_excel:
			from openpyxl import load_workbook, Workbook
			wb = load_workbook( filename=ops.targets_excel,  read_only=True)
			for row in wb[ wb.sheetnames[0] ].iter_rows( max_col=1):
				if not row[0].value.startswith('#'):
					inNetworks.add( row[0].value )
		else:
			print( 'Reading networks from STDIN:', file=sys.stderr)
			inNetworks = set(sys.stdin.read().splitlines())

		
		for nessus_file in args:
			with open( nessus_file, 'r' , encoding="utf8") as f:
				if ops.verbose:
					print( "Dumping data from: ", nessus_file, " - - - - - ") 
				if ops.dump_targets:
					for policy in xml.etree.ElementTree.parse(f).getroot().iter("preference"):
						if policy.find("name").text == "TARGET":
							for n in policy.find("value").text.split(','):
								nessusNetworks.add( n )
		print('Found in scans but not in STDIN/Excel:')
		for n in nessusNetworks.difference( inNetworks ):
			print( "'%s' as %s" %(n, ipOrName(n)))
		print('Found in STNIN/Excel but not in scans:')
		for n in inNetworks.difference( nessusNetworks ):
			print( "'%s' as %s" %(n, ipOrName(n)))
		exit()

	if ops.dump_xml_key:			
		for nessus_file in args:
			with open( nessus_file, 'r' , encoding="utf8") as f:
				print( "Dumping data from: ", nessus_file, " - - - - - ") 
				for policy in xml.etree.ElementTree.parse(f).getroot().iter(ops.dump_xml_key):			
					print( json.dumps( bf.data(policy), indent=4))
		exit()
	

	if ops.format in ('grep', 'xml'):
		outFile = None
		if ops.output_file:
			outFile = open( ops.output_file, 'w' , encoding="utf8")

	for nessus_file in args:
		report_task_id = False
		nameCache = {}
		with open( nessus_file, 'r', encoding="utf8") as f:
			if ops.verbose:
				print( "Reading data from %s " %nessus_file, file=sys.stderr)
			root = xml.etree.ElementTree.parse(f).getroot()
			report_task_id = getReportTaskId(root, nessus_file)
			try:
				for report in root.iter('Report'):
					for rh in report.findall("ReportHost"):
						cachedIP = False
						cachedLocation = False
						for ri in rh.findall("ReportItem"):
							# filter Severity
							if ops.min_severity <= int(ri.get("severity")):
								if ops.max_severity >= int(ri.get("severity")):
									# filter ID
									if ops.id == [] or str(ri.get("pluginID")) in ops.id:
										if ops.no_id == [] or ri.get("pluginID") not in ops.no_id:
											if not ops.match or ( ops.match and re.search(ops.match, getValue(ops.match_key, ri), re.IGNORECASE)):
												if not ops.no_match or (ops.no_match and not re.search(ops.no_match, getValue( ops.match_key, ri), re.IGNORECASE)):
													# because a ReportHost has multiple ReportItem, we need to cache this so it is only done once
													if not cachedIP:
														# get hostname so we can match host-filters with as well!
														cachedIP = True
														IP = tag(rh, 'host-ip')
														(hostName, hostNameProbability) = getBestHostname(ReportHost=rh, no_scanning= ops.lookups_no_certs, no_lookups=(not ops.lookups), resolvHelper=resolvHelper)
														if hostNameProbability.startswith('Correct'):
															displayAddr = hostName
														elif hostNameProbability.startswith('Guess'):
															displayAddr = "≈%s" % (hostName)
														else:
															displayAddr = ''
														
														# update cache
														if ops.lookups:
															nameCache[IP] = hostName, hostNameProbability
													
													# filter include or exclude this host?
													if includeHosts == set() or IP in includeHosts or hostName in includeHosts:
														if excludeHosts == set() or (IP not in excludeHosts and hostName not in excludeHosts):
															# filter done!

															if not cachedLocation:
																cachedLocation = True
																# do this after IP is matched
																ipLocation = ''
																if IP in hostLookup:
																	locationMatch = True
																	ipLocation = hostLookup[IP]
																if hostName in hostLookup:
																	if ipLocation:
																		ipLocation = "%s / %s" % ( ipLocation,hostLookup[hostName] )
																	else:
																		locationMatch = True
																		ipLocation = hostLookup[hostName]
																displayOs = getBestOS( ReportHost=rh )

															if ops.format in ['text', 'excel']:

																# one line per IP output
																long_findings.append( {"pluginID":ri.get("pluginID"), "risk_factor":ri.find("risk_factor").text, "pluginName":ri.get("pluginName"), "IP":addPort(IP, ri.get("port")), "hostname":addPort(displayAddr, ri.get("port")) ,'location':ipLocation, "severity":int(ri.get("severity")), "displayOs":displayOs})

																# compact output, meaning one file per finding
																if not ri.get("pluginID") in compact:
																	if displayAddr == '':
																		displayAddrCompact = IP
																	else:
																		displayAddrCompact = displayAddr

																	# new
																	findings.append( {"pluginID":ri.get("pluginID"), "severity":int(ri.get("severity")), "risk_factor":ri.find("risk_factor").text, "pluginName":ri.get("pluginName")})
																	compact[ri.get("pluginID")] = {"pluginID":ri.get("pluginID"), "risk_factor":ri.find("risk_factor").text, "pluginName":ri.get("pluginName"), "IP":[addPort(displayAddrCompact, ri.get("port"))], "severity":int(ri.get("severity"))}
																else:
																	# exists
																	compact[ri.get("pluginID")]["IP"].append( addPort(displayAddrCompact, ri.get("port")))

																# host stats 
																if ops.style == "host":
																	if IP not in host_findings:
																		host_findings[IP] = host_stats.copy()
																		host_findings[IP]["displayAddr"] = displayAddr
																		host_findings[IP]["location"] = ipLocation
																		host_findings[IP]["displayOs"] = displayOs
																		host_findings[IP]["hostNameProbability"] = hostNameProbability
																		
																	if int(ri.get("severity")) == severity["critical"]:
																		host_findings[IP]["critical"] = host_findings[IP]["critical"] + 1
																		sum_findings["critical"] = sum_findings["critical"] + 1
																	if int(ri.get("severity")) == severity["high"]:
																		host_findings[IP]["high"] = host_findings[IP]["high"] + 1
																		sum_findings["high"] = sum_findings["high"] + 1
																	if int(ri.get("severity")) == severity["medium"]:
																		host_findings[IP]["medium"] = host_findings[IP]["medium"] + 1
																		sum_findings["medium"] = sum_findings["medium"] + 1
																	if int(ri.get("severity")) == severity["low"]:
																		host_findings[IP]["low"] = host_findings[IP]["low"] + 1
																		sum_findings["low"] = sum_findings["low"] + 1
																	if int(ri.get("severity")) == severity["none"]:
																		host_findings[IP]["none"] = host_findings[IP]["none"] + 1
																		sum_findings["none"] = sum_findings["none"] + 1

															# html output
															elif ops.format in ['html', 'json']:
																if not ri.get("pluginID") in compact:
																	# first finding
																	findings.append( {"pluginID":ri.get("pluginID"), "severity":int(ri.get("severity")), "risk_factor":ri.find("risk_factor").text, "pluginName":ri.get("pluginName"), 'count':0})
																	compact[ri.get("pluginID")] = bf.data(ri)
																	compact[ri.get("pluginID")]["ReportItem"]["IP"] = [{"IP":IP, 'location':ipLocation, "port": ri.get("port"), "svc_name": ri.get("svc_name"), 'hostname':displayAddr, "hostNameProbability":hostNameProbability}]
																else:
																	# exists, add IP info only
																	compact[ri.get("pluginID")]["ReportItem"]["IP"].append( {"IP":IP, 'location':ipLocation, "port": ri.get("port"), "svc_name": ri.get("svc_name"), 'hostname':displayAddr,  "hostNameProbability":hostNameProbability} )

																# collect plugin output for all findings:
																if ri.find("plugin_output") != None and ri.find("plugin_output").text :
																	if (not "plugin_output" in compact[ri.get("pluginID")]["ReportItem"]) or type(compact[ri.get("pluginID")]["ReportItem"]["plugin_output"]) != list  :
																		# first plugin output for this finding
																		compact[ri.get("pluginID")]["ReportItem"]["plugin_output"] = [{"output":ri.find("plugin_output").text, "IP":IP, 'location':ipLocation, "port": ri.get("port"), "svc_name": ri.get("svc_name"), 'hostname':displayAddr}]
																	else:
																		# output exists, add data to that
																		compact[ri.get("pluginID")]["ReportItem"]["plugin_output"].append( {"output":ri.find("plugin_output").text, "IP":IP, 'location':ipLocation, "port": ri.get("port"), "svc_name": ri.get("svc_name"), 'hostname':displayAddr} )
																	# print( json.dumps(compact[ri.get("pluginID")]["ReportItem"]["plugin_output"] , indent=4))

															elif ops.format == 'xml':
																findings.append( ri )
																# xml output
																# only add newline when not already there
																nl = '\n'
																if ord(xml.etree.ElementTree.tostring(ri)[ -1:]) in [10, 13]:  # check last digit is 10 or 13. str compare failes on this bin object...
																	nl = ''
																	print(parseString( xml.etree.ElementTree.tostring(	ri )).toprettyxml(newl=nl), file=outFile)

															elif ops.format == 'grep':
																# <ReportItem pluginFamily="Windows" pluginID="11011" pluginName="Microsoft Windows SMB Service Detection" port="139" protocol="tcp" severity="0" svc_name="smb">
																# ...
																#         <plugin_output>
																elemList = []
																if not ops.grep_plugin:
																	elemList.append(ri.find("description"))
																if not ops.grep_description:
																	elemList.append(ri.find("plugin_output"))

																for po in elemList:
																	if type(po) == xml.etree.ElementTree.Element:
																		if po.text:
																			for l in po.text.splitlines():
																				if len(l) != 0:
																					if ops.grep_raw:
																						print("%s%s%s" % ( ops.grep_delim, l.strip(), ops.grep_delim), file=outFile)
																					else:
																						print( "%s\t%s\t%s%s%s" %(addPort(IP, ri.get("port")), addPort( displayAddr, ri.get("port")), ops.grep_delim, l.strip(), ops.grep_delim ), file=outFile)


			except (xml.etree.ElementTree.ParseError, UnicodeDecodeError) as err:
				print("Fatal: Failed to parse XML data: %s" %err,  file=sys.stderr)
				print("\n'%s' is most likely not a .nessus file!? \n" %nessus_file,  file=sys.stderr)
				exit()

		# save namecache
		if ops.lookups or ops.save_cache:
			with open( cachePath(report_task_id), 'w') as jf:
				json.dump( nameCache , jf)

	if ops.format in ['xml', 'grep']:
		if ops.output_file:
			close(outFile)
		exit()

	# sort data
	findings.sort(key=operator.itemgetter("severity", "pluginID"), reverse=True)
	long_findings.sort(key=operator.itemgetter("severity", "pluginID"), reverse=True)

	if ops.format in ('html', 'text', 'json', 'xml'):
		outFile = None
		if ops.output_file:
			outFile = open( ops.output_file, 'w' , encoding=sys.stdout.encoding)


	# text output
	if ops.format == 'text':
		pt = PrettyTableWrapper(vertical_char=ops.vertical_char)
		if ops.style == 'compact':
			pt.field_names = ["ID","severity","pluginName"]
			# build a table, then add addresses to end of each line since addresses can be both short and long
			addresses = []
			for f in findings:
				k = f.get("pluginID")
				pt.add_row([ compact[k].get("pluginID"), compact[k].get("risk_factor"), compact[k].get("pluginName")[:100] ])
				addresses.append( ",".join(compact[k].get("IP")) )
			s = ''
			for i, l in enumerate(pt.get_string().splitlines()):
				if i == 0: # first line is a header
					s = "%s Address-list\n" %l
				else:
					s += l + addresses[i-1] + '\n'
			print(s, file=outFile)

		if ops.style == 'long':
			if locationMatch:
				pt.field_names = ["ID","severity","pluginName","hostname", "IP", "Location", "Operative-System"]
				for f in long_findings:
					pt.add_row( [f.get("pluginID"), f.get("risk_factor"), f.get("pluginName")[:55], f.get("hostname"), f.get("IP"), f.get("location"), f.get("displayOs") ])
			else:
				pt.field_names = ["ID", "severity", "pluginName", "hostname", "IP", "Operative-System"]
				for f in long_findings:
					pt.add_row( [f.get("pluginID"), f.get("risk_factor"), f.get("pluginName")[:55], f.get("hostname"), f.get("IP"), f.get("displayOs")])
			print(pt, file=outFile)
		if ops.style == 'host':
			if locationMatch:
				pt.field_names = ["Critical", "High", "Medium", "Low", "None", "Addr", "IP", "Location", "Operative-System", "hostname-probability"]
				for k in host_findings.keys():
					pt.add_row( [ host_findings[k]["critical"], host_findings[k]["high"], host_findings[k]["medium"], host_findings[k]["low"], host_findings[k]["none"], host_findings[k]["displayAddr"], k, host_findings[k]["location"][:50], host_findings[k]["displayOs"][:50], host_findings[k]["hostNameProbability"] ] )
			else:
				pt.field_names = ["Critical", "High", "Medium", "Low", "None", "Addr", "IP", "Operative-System", "hostname-probability"]
				for k in host_findings.keys():
					pt.add_row( [ host_findings[k]["critical"], host_findings[k]["high"], host_findings[k]["medium"], host_findings[k]["low"], host_findings[k]["none"], host_findings[k]["displayAddr"], k, host_findings[k]["displayOs"][:50], host_findings[k]["hostNameProbability"] ] )

			print(pt, file=outFile)

			print('---------------------- Summary of findings from above: -------------------------', file=outFile)

			ptSum = PrettyTableWrapper(vertical_char=ops.vertical_char)
			ptSum.field_names = ["Critic", "High", "Medium", "Low", "None"]
			ptSum.add_row( [sum_findings['critical'], sum_findings['high'], sum_findings['medium'], sum_findings['low'], sum_findings['none']])
			print(ptSum, file=outFile)

			print("Total number of hosts:\t", len(host_findings), sep="\t", file=outFile)
			uncompliant_cnt=0
			for k in host_findings.keys():
				if host_findings[k]["critical"] > 0 or host_findings[k]["high"] > 0:
					uncompliant_cnt = uncompliant_cnt + 1
			try:
				p = format(round(uncompliant_cnt / len(host_findings)*100))
			except ZeroDivisionError:
				p = 0

			print("Uncompliant number of hosts:", uncompliant_cnt,  '( %s%% )' %p , sep="\t", file=outFile)
			print('-'*80, file=outFile)
				
	# html output
	elif ops.format in ['html', 'json']:
		# custom template
		if ops.html_template:
			with open(ops.html_template, "r", encoding="utf8") as t:
				mustache_template = t.read()
			
		details = []
		for k in compact.keys():
			details.append( compact[ k ]["ReportItem"] )

		# add finding count to Table of findings:
		for f in findings:
			f['count'] = len( compact[f["pluginID"]]["ReportItem"]["IP"])

		# render
		jReports = {"details":details, "severityList":findings }
		if ops.format == 'json':
			print( json.dumps(jReports, indent=4), file=outFile)
		else:
			print( pymustache.render( mustache_template, jReports ), file=outFile)
	
	# excel output 
	elif ops.format == 'excel':
		excel_wb = None
		excel_long = None
		excel_compact = None
		
		if ops.style == "host":
			if os.path.isfile( ops.output_file ):
				import zipfile
				try:
					excel_wb = load_workbook( filename = ops.output_file)
				except zipfile.BadZipFile:
					print("Fatal: the --output-file '%s' exists, but is not an excel file?" %ops.output_file, file=sys.stderr )
					exit()
				excel_s = excel_wb.create_sheet("hosts_" + ops.output_sheet)
			else:
				excel_wb = Workbook()
				excel_s = excel_wb.active
				excel_s.title = "hosts_" + ops.output_sheet
			excel_s.append(("Critical", "High", "Medium", "Low", "None", "Compliant", "Address", "IP", "Location", "Operative-System", "hostname-probability"))
			i=0
			for i, k in enumerate(host_findings.keys()):
				excel_s.append((int(host_findings[k]["critical"]), int(host_findings[k]["high"]), int(host_findings[k]["medium"]), int(host_findings[k]["low"]), int(host_findings[k]["none"]), "=NOT(OR(A" + str(i + 2) + ":B" + str(i + 2) + "))", host_findings[k]["displayAddr"], k, host_findings[k]["location"],
				host_findings[k]["displayOs"], host_findings[k]["hostNameProbability"]))
			excel_s.append(( "=AVERAGE(A2:A"+str(i+2)+")", "=AVERAGE(B2:B"+str(i+2)+")", "=AVERAGE(C2:C"+str(i+2)+")", "=AVERAGE(D2:D"+str(i+2)+")", "=AVERAGE(E2:E"+str(i+2)+")", "=ROUND(AVERAGEA(F2:F"+str(i+2)+")*100,0) &\"% compliance level\"" ))

		else:
			# add to existing book 
			if os.path.isfile( ops.output_file ):
				excel_wb = load_workbook( filename = ops.output_file)
				excel_long = excel_wb.create_sheet("long_" + ops.output_sheet)
			else:
				# new book 
				excel_wb = Workbook()
				excel_long = excel_wb.active
				excel_long.title = "long_" + ops.output_sheet
			excel_compact = excel_wb.create_sheet("Compact_" + ops.output_sheet)
			
			excel_compact.append( ("ID","Severity","pluginName","Address" ))
			excel_long.append(( "ID","Severity","pluginName","Hostname", "IP", "Location", "Operative system", "Remediation status", "Owner" ))

			for f in findings:
				k = f.get("pluginID")
				excel_compact.append( ( int(compact[k].get("pluginID")), compact[k].get("risk_factor"), compact[k].get("pluginName"), ",".join(compact[k].get("IP"))) ) 

			for f in long_findings:
				excel_long.append(( int(f.get("pluginID")), f.get("risk_factor"), f.get("pluginName"), f.get("hostname"), f.get("IP"), f.get("location"),f.get("displayOs")))
	
		excel_wb.save( ops.output_file )
		
	
		
		